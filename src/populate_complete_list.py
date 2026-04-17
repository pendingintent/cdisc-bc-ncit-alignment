"""Populate output/Complete List.xlsx from the NCIt EVS REST API.

Pipeline:
  1. Load the column headers from files/Complete List.xlsx (Sheet1, row 1).
  2. Call GET /concept/ncit/codes to get the full NCIt code list.
  3. Fetch concept details in concurrent batches and map each to a row.
  4. Write a fresh workbook to output/Complete List.xlsx using the template's headers.

Writing uses openpyxl's write_only mode — the ~212k row workbook would otherwise
balloon memory. Rows are streamed in as each batch arrives.

Usage:
    python -m src.populate_complete_list
    python -m src.populate_complete_list --limit 500    # smoke test
    python -m src.populate_complete_list --batch-size 250 --workers 12
"""

from __future__ import annotations

import argparse
import pathlib
import sys
import time

from openpyxl import Workbook, load_workbook

from src.cli_utils import validate_xlsx_path, xlsx_path
from src.ncit_client import fetch_all_codes, fetch_concepts, make_session
from src.ncit_mapping import concept_to_row

ROOT = pathlib.Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "files" / "Complete List.xlsx"
OUTPUT = ROOT / "output" / "Complete List.xlsx"
SHEET_NAME = "Sheet1"


def read_headers(template: pathlib.Path) -> list[str]:
    """Return the header row (row 1) of the template Complete List workbook."""
    wb = load_workbook(template, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    finally:
        wb.close()


def _log_progress(done: int, total: int, start: float) -> None:
    """Emit a one-line progress update to stderr with running rate and ETA."""
    elapsed = time.monotonic() - start
    rate = done / elapsed if elapsed else 0
    remaining = (total - done) / rate if rate else 0
    print(
        f"  batches {done}/{total}  elapsed={elapsed:6.1f}s  "
        f"rate={rate:5.2f} b/s  eta={remaining:6.1f}s",
        file=sys.stderr,
        flush=True,
    )


def populate(
    output_path: pathlib.Path = OUTPUT,
    template_path: pathlib.Path = TEMPLATE,
    batch_size: int = 200,
    workers: int = 8,
    limit: int | None = None,
) -> pathlib.Path:
    """Fetch every NCIt concept and stream-write the Complete List workbook."""
    validate_xlsx_path(output_path, label="--output")
    validate_xlsx_path(template_path, label="--template")
    headers = read_headers(template_path)
    session = make_session(pool=workers * 2)

    print("Fetching NCIt code list...", file=sys.stderr, flush=True)
    codes = fetch_all_codes(session=session)
    if limit is not None:
        codes = codes[:limit]
    print(f"  {len(codes):,} codes", file=sys.stderr, flush=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # write_only keeps memory flat while streaming ~212k rows.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=SHEET_NAME)
    ws.append(headers)

    start = time.monotonic()
    written = 0
    for concept in fetch_concepts(
        codes,
        session=session,
        batch_size=batch_size,
        max_workers=workers,
        on_progress=lambda d, t: _log_progress(d, t, start),
    ):
        ws.append(concept_to_row(concept))
        written += 1

    print(f"Writing {written:,} rows to {output_path}", file=sys.stderr, flush=True)
    wb.save(output_path)
    print(f"Done in {time.monotonic() - start:.1f}s", file=sys.stderr, flush=True)
    return output_path


def main(argv: list[str] | None = None) -> int:
    """CLI entry point; parses arguments and delegates to :func:`populate`."""
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", type=xlsx_path("--output"), default=OUTPUT, help="output .xlsx path")
    parser.add_argument("--template", type=xlsx_path("--template"), default=TEMPLATE, help="template .xlsx path")
    parser.add_argument("--batch-size", type=int, default=200)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--limit", type=int, default=None, help="cap number of codes (smoke test)")
    args = parser.parse_args(argv)

    populate(
        output_path=args.output,
        template_path=args.template,
        batch_size=args.batch_size,
        workers=args.workers,
        limit=args.limit,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
