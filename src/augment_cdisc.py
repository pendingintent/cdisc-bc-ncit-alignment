"""Augment Complete List.xlsx with CDISC BC lookup columns.

Appends three columns to each existing row:
  - Exists in CDISC   (True/False)
  - CDISC href        (/mdr/bc/biomedicalconcepts/<code>)
  - CDISC title       (BC label)

Values come from a single call to the CDISC Library COSMOS endpoint
`/mdr/bc/biomedicalconcepts`. The NCIt code in the first column is matched
against the final path segment of each returned href.

Both --input and --output must end with .xlsx. If --output equals --input the
update is in place: rows are streamed from the input, a temp file is written
next to it, and the original is atomically replaced.

Usage:
    python -m src.augment_cdisc                                   # defaults: inline on output/Complete List.xlsx
    python -m src.augment_cdisc --input in.xlsx --output out.xlsx
"""

from __future__ import annotations

import argparse
import os
import pathlib
import sys
import tempfile
import time

from openpyxl import Workbook, load_workbook

from src.cdisc_client import fetch_biomedical_concepts, make_session
from src.cdisc_mapping import build_index, lookup
from src.cli_utils import validate_xlsx_path, xlsx_path

ROOT = pathlib.Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "output" / "Complete List.xlsx"
SHEET_NAME = "Sheet1"

NEW_HEADERS = ["Exists in CDISC", "CDISC href", "CDISC title"]


def augment(
    input_path: pathlib.Path = DEFAULT_XLSX,
    output_path: pathlib.Path | None = None,
) -> pathlib.Path:
    """Stream the input workbook and append CDISC lookup columns to each row."""
    validate_xlsx_path(input_path, label="--input")
    if output_path is None:
        output_path = input_path
    validate_xlsx_path(output_path, label="--output")

    if not input_path.exists():
        raise FileNotFoundError(input_path)

    print("Fetching CDISC biomedical concepts index...", file=sys.stderr, flush=True)
    session = make_session()
    response = fetch_biomedical_concepts(session=session)
    index = build_index(response)
    print(f"  {len(index):,} CDISC BC entries indexed", file=sys.stderr, flush=True)

    start = time.monotonic()
    inline = input_path.resolve() == output_path.resolve()

    # Stream rows from the input workbook into a write-only workbook so the full
    # 212k-row sheet never lives in memory at once.
    src_wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        src_ws = src_wb[SHEET_NAME]
        out_wb = Workbook(write_only=True)
        out_ws = out_wb.create_sheet(title=SHEET_NAME)

        total = 0
        hits = 0
        for idx, row in enumerate(src_ws.iter_rows(values_only=True)):
            row_list = list(row)
            if idx == 0:
                out_ws.append(row_list + NEW_HEADERS)
                continue
            code = row_list[0]
            exists, href, title = lookup(index, code) if code else (False, "", "")
            if exists:
                hits += 1
            out_ws.append(row_list + [exists, href, title])
            total += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)

        if inline:
            # Write to a sibling temp file, then atomically replace the input so a
            # mid-write failure cannot corrupt the source workbook.
            tmp_fd, tmp_name = tempfile.mkstemp(
                prefix=output_path.stem + ".", suffix=".xlsx.tmp", dir=output_path.parent
            )
            os.close(tmp_fd)
            tmp_path = pathlib.Path(tmp_name)
            try:
                out_wb.save(tmp_path)
                os.replace(tmp_path, output_path)
            finally:
                if tmp_path.exists():
                    tmp_path.unlink()
        else:
            out_wb.save(output_path)
    finally:
        src_wb.close()

    elapsed = time.monotonic() - start
    print(
        f"Wrote {total:,} rows ({hits:,} CDISC hits) to {output_path} — {elapsed:.1f}s",
        file=sys.stderr,
        flush=True,
    )
    return output_path


def main(argv: list[str] | None = None) -> int:
    """CLI entry point; parses arguments and delegates to :func:`augment`."""
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--input",
        type=xlsx_path("--input"),
        default=DEFAULT_XLSX,
        help="source .xlsx produced by populate_complete_list",
    )
    parser.add_argument(
        "--output",
        type=xlsx_path("--output"),
        default=None,
        help="destination .xlsx (defaults to --input for in-place update)",
    )
    args = parser.parse_args(argv)
    augment(args.input, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
